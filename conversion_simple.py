from openpyxl import load_workbook

# Chemin vers ton fichier Excel
fichier = r"C:\Users\ychaker\Desktop\Mesures_Compensatoires_Output\mesures_compensatoires_PK_20250715_1354.xlsx"

# Charger le fichier Excel avec styles
wb = load_workbook(fichier)
ws = wb.active  # ou ws = wb['NomFeuille'] si tu connais le nom exact

# Trouver les indices des colonnes PK_D et PK_F
header = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
if 'PK_D' not in header or 'PK_F' not in header:
    raise ValueError("Les colonnes 'PK_D' et 'PK_F' doivent être présentes dans la première ligne.")

col_pk_d = header['PK_D']
col_pk_f = header['PK_F']

# Parcourir les lignes et inverser si PK_D > PK_F
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    cell_d = row[col_pk_d - 1]
    cell_f = row[col_pk_f - 1]

    # Vérifier que les cellules contiennent bien des nombres
    try:
        val_d = float(cell_d.value)
        val_f = float(cell_f.value)
    except (TypeError, ValueError):
        continue  # ignorer les lignes avec des valeurs non numériques

    if val_d > val_f:
        # Inverser les valeurs (tout en gardant la mise en forme)
        cell_d.value, cell_f.value = val_f, val_d

# Enregistrer dans un nouveau fichier
fichier_corrige = fichier.replace(".xlsx", "_corrigé.xlsx")
wb.save(fichier_corrige)

print(f"✅ Fichier corrigé (avec style conservé) enregistré ici : {fichier_corrige}")
