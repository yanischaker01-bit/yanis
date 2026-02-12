import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# ---------------------------------------------------------
# PARAMÈTRES
# ---------------------------------------------------------
FICHIER = r"C:\Users\ychaker\Desktop\segmentation 2021 INGESCAN\fusion_ods_formaté_par_photo.xlsx"

# ---------------------------------------------------------
# 0) IDENTIFICATION AUTOMATIQUE DE LA BONNE FEUILLE
# ---------------------------------------------------------
xls = pd.ExcelFile(FICHIER)

colonnes_requises = {"libelle_objet", "pos_dpf", "code_cosea", "Source_Feuille", "longueur"}
feuille_cible = None

for sheet in xls.sheet_names:
    df_test = pd.read_excel(FICHIER, sheet_name=sheet)
    if colonnes_requises.issubset(df_test.columns):
        feuille_cible = sheet
        break

if feuille_cible is None:
    raise Exception(
        "❌ Aucune feuille ne contient toutes les colonnes requises : "
        + ", ".join(colonnes_requises)
    )

print(f"✔ Feuille détectée automatiquement : {feuille_cible}")

# ---------------------------------------------------------
# Chargement de la feuille correcte
# ---------------------------------------------------------
df = pd.read_excel(FICHIER, sheet_name=feuille_cible, dtype=str)
df["longueur"] = pd.to_numeric(df["longueur"], errors="coerce").fillna(0)

df["libelle_objet"] = df["libelle_objet"].fillna("Sans_libelle")
df["pos_dpf"] = df["pos_dpf"].fillna("Sans_position")
df["code_cosea"] = df["code_cosea"].fillna("Sans_code")
df["Source_Feuille"] = df["Source_Feuille"].fillna("Inconnue")

# ---------------------------------------------------------
# Styles (identiques à tes feuilles existantes)
# ---------------------------------------------------------
thin = Side(border_style="thin", color="000000")
border_all = Border(top=thin, left=thin, right=thin, bottom=thin)
header_fill = PatternFill("solid", fgColor="BDD7EE")
align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

def formater_colonnes(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

def mettre_en_forme(ws, nb_cols):
    for row in ws.iter_rows(min_row=1, max_col=nb_cols):
        for cell in row:
            cell.border = border_all
            cell.alignment = align_center

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True)

# ---------------------------------------------------------
# 1) Résumé_length
# ---------------------------------------------------------
res = (
    df.groupby(["code_cosea", "libelle_objet", "pos_dpf"])["longueur"]
    .sum()
    .reset_index()
)

totaux = df.groupby("code_cosea")["longueur"].sum().reset_index()
totaux["libelle_objet"] = "TOTAL"
totaux["pos_dpf"] = ""

res_complet = pd.concat([res, totaux], ignore_index=True)

# ---------------------------------------------------------
# 2) Résumé_Source_Feuille_length
# ---------------------------------------------------------
res_src = (
    df.groupby(["Source_Feuille", "libelle_objet", "pos_dpf"])["longueur"]
    .sum()
    .reset_index()
)

totaux_src = df.groupby("Source_Feuille")["longueur"].sum().reset_index()
totaux_src["libelle_objet"] = "TOTAL"
totaux_src["pos_dpf"] = ""

res_src_complet = pd.concat([res_src, totaux_src], ignore_index=True)

# ---------------------------------------------------------
# 3) Stats_Code_Cosea_length
# ---------------------------------------------------------
stats_code = df.groupby("code_cosea")["longueur"].sum().reset_index()

# ---------------------------------------------------------
# 4) Stats_Source_Feuille_length
# ---------------------------------------------------------
stats_src = df.groupby("Source_Feuille")["longueur"].sum().reset_index()

# ---------------------------------------------------------
# Ajout des nouvelles feuilles SANS toucher aux existantes
# ---------------------------------------------------------
wb = openpyxl.load_workbook(FICHIER)

def ajouter_feuille(nom, dataframe):
    ws = wb.create_sheet(nom)
    for r in dataframe_to_rows(dataframe, index=False, header=True):
        ws.append(r)
    mettre_en_forme(ws, len(dataframe.columns))
    formater_colonnes(ws)

ajouter_feuille("Résumé_length", res_complet)
ajouter_feuille("Résumé_Source_Feuille_length", res_src_complet)
ajouter_feuille("Stats_Code_Cosea_length", stats_code)
ajouter_feuille("Stats_Source_Feuille_length", stats_src)

wb.save(FICHIER)

print("✔ Feuilles ajoutées avec succès (aucune feuille modifiée).")
